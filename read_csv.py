from os import listdir
from os.path import isfile, join
from datetime import datetime
from openpyxl import Workbook
import csv
import os


def readcsv():
    def extract_month_from_date(date_string):
        try:
            # example for dateformat: "29.03.2023 11:21"
            date_obj = datetime.strptime(date_string, "%d.%m.%Y %H:%M")
            return date_obj.strftime("%m-%Y")  # return month and year
        except ValueError:
            return None  # return None if date not valid

    current_path = os.getcwd()
    current_path = current_path + r"\csv_file"
    nur_dateline = [f for f in listdir(current_path) if isfile(join(current_path, f))]
    date = fr"{current_path}\{nur_dateline[0]}"

    results = {}
    # Initialise a dictionary to sava the results 
    with open(date, "r") as date:
        csv_reader = csv.reader(date, delimiter=";")
        next(csv_reader)  # skip header

        for line in csv_reader:
            name = line[0]  # name from first column
            datum = line[2]  # date from third column
            column6_value = int(line[5])  # Value convert to int from column 6

            # Extract the month from date
            month = extract_month_from_date(datum)

            # only if date is fine. it will continue 
            if month is not None:
                # Create a unique key name + month
                entry = f"{name}_{month}"

                # if entry exists in dictionary adds the values
                if entry in results:
                    results[entry] += column6_value
                else:
                    results[entry] = column6_value
    # Create Excel file and write the data in tables 
    wb = Workbook()

    # a table tab will create for every year for the data block
    for entry, total in results.items():
        name, month = entry.split('_')
        sign = name

        year = month.split('-')[1]  # extract the year from month
        month_num = int(month.split('-')[0])  # convert the month in an int (1- 12)

        # if table tab for this year doesn't exist, create it
        if year not in wb.sheetnames:
            wb.create_sheet(title=year)

        sheet = wb[year]

        # sign is write in column A
        if sign not in [cell.value for cell in sheet['A']]:  # List the Value at column A
            sheet.append([sign])

        # if month not exist in the header, it will be created
        if month_num not in range(1, 13):
            continue

        month_column_index = month_num + 1  # column index +1 because month header

        if sheet.cell(row=1, column=month_column_index).value is None:
            sheet.cell(row=1, column=month_column_index).value = month

        # total way entry
        sign_row = [cell.value for cell in sheet['A']].index(sign) + 1  # Row index for sign
        sheet.cell(row=sign_row, column=month_column_index).value = total

    # Sava Excel-file
    excel_file = "Fahrzeugdaten.xlsx"  # File name for Excel
    wb.remove(wb['Sheet'])
    wb.save(f"out/{excel_file}")
    os.remove(fr"{current_path}\{nur_dateline[0]}")
